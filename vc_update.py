from os import getcwd
from time import sleep
from url_storage import provide_vcommerce_url
import os
from openpyxl import load_workbook
import datetime
import clipboard
from map import Map
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service as ChromeService
from subprocess import CREATE_NO_WINDOW

class Updater:

    def __init__(self, vc_password, update_count):
        self.vc_password = vc_password
        self.update_count = int(update_count)
        self.cwd = getcwd()
        self.ex_workbook, self.ex_worksheet = self.create_excel_update()
        self.driver = self.chromedriver_setup()
        self.actions = ActionChains(self.driver)
        self.map = Map()
        self.supplier_id = ''
        self.vc_id = ''
        self.vendor_name = ''
        self.counter = 1    

    # Sets up the chromedriver as part of the constructor. Requires a copy of Chrome's user data folder to be located
    # in the working directory, as well as a version of chromedriver.exe that matches your web browser version.
    def chromedriver_setup(self):
        print('Initializing Webdriver')
        driver_options = webdriver.ChromeOptions()
        driver_options.add_argument('user-data-dir=' + self.cwd + '/User Data')
        chrome_service = ChromeService(self.cwd + '/chromedriver.exe')
        chrome_service.creationflags = CREATE_NO_WINDOW
        create_driver = webdriver.Chrome(options=driver_options, executable_path=self.cwd + '/chromedriver.exe',
                                         service=chrome_service)
        return create_driver

    # Closes down all pages on the chromedriver.
    def chromedriver_close(self):
        self.driver.quit()

    # Opens and returns the Excel spreadsheet for the day. If that days' spreadsheet does not exist, it
    # will create one based off the template found in the working directory.
    @staticmethod
    def create_excel_update():
        downloads_path = os.path.expanduser("~") + "/Downloads/"
        try:
            ex_workbook = load_workbook(downloads_path + "BCS_New_Vendors_" + str(datetime.date.today()) + ".xlsx")
        except FileNotFoundError:
            ex_workbook = load_workbook("BCS_template.xlsx")
        ex_worksheet = ex_workbook.worksheets[1]
        return ex_workbook, ex_worksheet

        # Saves all changes made to the Excel spreadsheet and adjust the length of cells so all information is readable.
    # If this is a newly created spreadsheet it will be named: "BCS_New_vendors_<today's_date>"
    def format_and_save_excel(self):
        column_names = ('A', 'B', 'C')
        downloads_path = os.path.expanduser("~") + "/Downloads/"
        index_counter = 0
        max_length = 0
        for column in self.ex_worksheet.columns:
            length_list = []
            for cell in column:
                length = len(str(cell.value))
                length_list.append(length)
                max_length = max(length_list)
            self.ex_worksheet.column_dimensions[column_names[index_counter]].width = (max_length + 2)
            index_counter += 1
        self.ex_workbook.save(downloads_path + 'BCS_New_Vendors_' + str(datetime.date.today())
                              + '.xlsx')

    # Adds the provided value parameter to the spreadsheet cell located at the provided cell_coord parameter.
    def add_to_excel(self, cell_coord, value):
        cell = self.ex_worksheet[cell_coord]
        cell.value = value

    # Iterates through the Excel spreadsheet to find and return the coordinates of the first blank cell in the
    # column identified by the target_column parameter.
    def find_blank_cell_in_column(self, target_column):
        column_coord = target_column[0].coordinate
        for row in self.ex_worksheet.iter_rows(1, 100):
            for cell in row:
                if cell.coordinate[0] == column_coord[0]:
                    if cell.value is None:
                        return cell.coordinate
                    else:
                        continue
                else:
                    continue

    def update_spreadsheet(self):
        vendor_name_cell = self.find_blank_cell_in_column(self.ex_worksheet['B'])
        self.add_to_excel(vendor_name_cell, self.vendor_name)
        vc_num_cell = self.find_blank_cell_in_column(self.ex_worksheet['A'])
        self.add_to_excel(vc_num_cell, self.vc_id)
        supplier_id_cell = self.find_blank_cell_in_column(self.ex_worksheet['C'])
        self.add_to_excel(supplier_id_cell, self.supplier_id)

    # The clipboard_copy() method copies any arguments provided into the clipboard. This is useful for quickly getting
    # the entire text from a website.
    @staticmethod
    def clipboard_copy(text):
        results = clipboard.copy(text)
        return results

    @staticmethod
    def clipboard_paste():
        results = clipboard.paste()
        return results

    # The switch_tab_to_1() method has the driver object switch to the first browser tab, which has Vcommerce loaded.
    def switch_tab_to_1(self):
        self.driver.switch_to.window(self.driver.window_handles[0])

    # The switch_tab_to_2() method has the driver object switch to the second browser tab, which has Riskonnect loaded.
    def switch_tab_to_2(self):
        self.driver.switch_to.window(self.driver.window_handles[1])

    # The close_current_tab method closes the currently open tab.
    def close_current_tab(self):
        self.driver.close()

    # The open_vcommerce() method has the webdriver object open the Vendor ID update list in Vcommerce and then wait
    # for the list to load.
    def open_vcommerce(self):
        print('Opening Vcommerce')
        try:
            self.driver.get(provide_vcommerce_url())
        except NameError:
            self.driver.get()
            raise Exception("Error, the vcommerce URL was not provided. This URL is returned by the provide_vcommerce_url() function. This is located in url_storage.py, which has not been uploaded to Github due to company privacy reasons")
        sleep(3)

        # Sometimes Vcommerce will go directly to the page without going to the login screen first. The following
        # logic checks to see if the login page has loaded, and will log in if so. If the login page has not loaded
        # it will check to see if the Vcommerce search results have loaded and continue on if so (and raise a timeout
        # exception if not).
        login_verify = self.driver.find_elements(By.XPATH, self.map.full_xpath['approve_login_button'])
        if login_verify:
            vc_password_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_vc_password'])
            vc_password_element.send_keys(' ')
            vc_password_element.clear()
            sleep(0.5)
            vc_password_element.send_keys(self.vc_password)
            sleep(0.5)
            login_button_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_login_button'])
            login_button_element.click()
            WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH, self.map.full_xpath[
                'approve_search_results'])))
        else:
            WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH, self.map.full_xpath[
                'approve_search_results'])))

    # The vc_open_summary() method has the webdriver select and open the first vendor link on the Vcommerce
    # vendor list in a new tab.
    def vc_open_summary(self):
        print(f'Accessing vendor {self.counter}')
        item_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_first_item'])
        self.actions.key_down(Keys.CONTROL).perform()
        self.actions.click(item_element).perform()
        self.actions.key_up(Keys.CONTROL).perform()

    def vc_access_summary(self):
        vc_summary_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_summary_button'])
        vc_summary_element.click()
        WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH, self.map.full_xpath['approve_summary_page'
                                                                                                         '_body'])))
        vc_summary_body_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_summary_page_body'])
        vc_summary_body_element.send_keys(Keys.CONTROL + 'a')
        vc_summary_body_element.send_keys(Keys.CONTROL + 'c')

    # The parse_summary() method takes the copied data from the vc_summary and locates the supplier ID number and
    # Vcommerce ID numbers, which are then saved as class attributes.
    def parse_summary(self):
        # Takes the clipboard data as a giant string
        summary_text = self.clipboard_paste()
        # creates two lists from the data by splitting on whitespace
        summary_list = summary_text.split()
        summary_list_2 = summary_text.split()
        # looks through the first list for the string 'Number'. When found it saves the index location and checks if the
        # list item prior to the saved index is the string 'Supplier'. If so, the list item after the saved index
        # is saved. If not, it deletes that instance of the string 'Number' from the list and looks for the next.
        # this loops until it locates the correct combination of strings 'Supplier' and 'Number'
        for item in summary_list:
            if item == 'Number':
                temp_index = summary_list.index(item)
                if summary_list[temp_index - 1] == 'Supplier':
                    self.supplier_id = summary_list[temp_index + 1]
                    break
                else:
                    summary_list.remove(item)
                    continue
        # looks through the second list for the string 'ID'. When found the string index position in the list is saved
        # then the next item in the list (saved index + 1) is then saved as a class attribute. When this is found
        # the loop is broken.
        for item in summary_list_2:
            if item == 'ID':
                temp_index = summary_list_2.index(item)
                self.vc_id = summary_list_2[temp_index + 1]
                break
        self.find_vendor_name(summary_text)

    def find_vendor_name(self, input_string):
        input_list_split = input_string.split('\n')
        for item in input_list_split:
            if item == 'Registration Status\r':
                temp_index = input_list_split.index(item)
                if input_list_split[temp_index - 2] == 'Supplier Number\r':
                    vendor_name_unformatted = input_list_split[temp_index - 4]
                    self.vendor_name = self.vendor_name_format(vendor_name_unformatted)
                    sleep(0.5)
                else:
                    vendor_name_unformatted = input_list_split[temp_index - 2]
                    self.vendor_name = self.vendor_name_format(vendor_name_unformatted)
                    sleep(0.5)

    @staticmethod
    def vendor_name_format(vendor_name_unformatted_arg):
        if vendor_name_unformatted_arg[0:20] == 'Doing Business As : ':
            temp_vendor_name = vendor_name_unformatted_arg[20:]
            vendor_name_formatted1 = temp_vendor_name.title()
        else:
            vendor_name_formatted1 = vendor_name_unformatted_arg.title()
        if (vendor_name_formatted1[-4:]) == 'Llc\r':
            vendor_name_formatted2 = (vendor_name_formatted1[:-4] + 'LLC')
        else:
            vendor_name_formatted2 = vendor_name_formatted1[:-1]
        vendor_name_split = vendor_name_formatted2.split()
        vendor_name_formatted3 = ''
        for item in vendor_name_split:
            if item not in ['Of', 'Is', 'The']:
                vendor_name_formatted3 += item + ' '
            else:
                vendor_name_formatted3 += item.lower() + ' '
        if vendor_name_formatted3[0:3] == 'the':
            vendor_name_formatted3 = 'The' + vendor_name_formatted3[3:]
        vendor_name_formatted3 = vendor_name_formatted3[:-1]
        if vendor_name_formatted3[-3:] == 'Inc':
            vendor_name_formatted3 = vendor_name_formatted3 + '.'
        return vendor_name_formatted3

    # The vc_remove_item() method has the driver click the 'remove item' button on the vendor list once the update
    # is completed.
    def vc_remove_item(self):
        print('Removing vendor')
        remove_button_element = self.driver.find_element(By.XPATH, self.map.full_xpath['approve_remove_button'])
        remove_button_element.click()

    # The wait_for_remove() method has the driver wait until the item has been removed from the Vcommerce list
    # before moving forward.
    def wait_for_remove(self):
        sleep(2)
        WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH, self.map.full_xpath['approve_remove_'
                                                                                                         'button'])))
        print(f'Vendor {self.counter} updated\n')
        self.counter += 1

    # The prep_update() method runs both the open_vcommerce() method and the open_riskonnect() methods.
    # This is for use in the __main__ script and GUI.
    def prep_update(self):
        self.open_vcommerce()

    # The process_update() method runs all methods in the order required to process the vendor update. This is for
    # use in the __main__ script and GUI.
    def process_update(self):
        for i in range(self.update_count):
            self.switch_tab_to_1()
            self.vc_open_summary()
            self.switch_tab_to_2()
            self.vc_access_summary()
            self.close_current_tab()
            self.switch_tab_to_1()
            self.parse_summary()
            self.update_spreadsheet()
            self.format_and_save_excel()
            self.vc_remove_item()
            self.wait_for_remove()
