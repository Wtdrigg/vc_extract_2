"""
This module contains the Extractor class and imports all of its dependencies. This is instantiated when using
the "New Extract" or "Existing Extract" buttons in the GUI.

IMPORTANT - I made several edits to this in my last few days at vulcan in an attempt to have this
work for other users. As of the time of writing this appears to work as desired but some of the code
changes were a bit messy due to the time constraint (in particular the use of HTML paths here and in
the map.py file.) If I had more time I would clean this up more but it works for now, so I apologize for 
the mess.
"""

import os
import time
import datetime
import clipboard
import pickle
import traceback
from openpyxl import load_workbook
from time import sleep
from map import Map
from vendor_info import VendorInfo
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.service import Service as ChromeService
from subprocess import CREATE_NO_WINDOW


# Extractor class is used to pull all vendor info from the Vcommerce system and add it to an Excel spreadsheet.
# All vendor info pulled on the same day will go to the same Excel file.
class Extract:

    # Constructor sets class attributes, opens the Excel file, and creates objects that will be used later.
    def __init__(self, vc_password, vc_number):
        try:
            self.vc_password = vc_password
            self.vc_number = vc_number
            self.vc_service_type = ''
            self.input_string = ''
            self.input_list = []
            self.cwd = os.getcwd()
            self.driver = self.chromedriver_setup()
            self.map = Map()
            self.actions = ActionChains(self.driver)
            self.ex_workbook, self.ex_worksheet = (None, None)
            self.vendor_info = self.load_vendor_info()
        except:
            print(traceback.format_exc())

    def load_vendor_info(self):
        try:
            with open('J:/VC_Extract_Files/vendor_info.pkl', 'rb') as file:
                vendor_info_object = pickle.load(file)
                return vendor_info_object
        except FileNotFoundError:
            vendor_info_object = self.save_vendor_info()
            return vendor_info_object
        
    def save_vendor_info(self):
        vi = VendorInfo()
        with open('J:/VC_Extract_Files/vendor_info.pkl', 'wb') as file:
            pickle.dump(vi, file)
        return vi
        

    # Checks if the input provided by the users is a correct 10 digit Vcommerce number.
    def verify_user_input(self):
        print('Verifying correct user input')
        user_input = self.vc_number
        if len(user_input) == 10:
            is_ten_digits = True
        else:
            is_ten_digits = False
        if user_input.isdigit():
            is_numbers = True
        else:
            is_numbers = False
        if is_ten_digits is False or is_numbers is False:
            raise ValueError('Your Entry did not contain a valid Vcommerce number, please verify that you did '
                             'not \n accidentally copy any whitespace before the number as that is the most common '
                             'cause of this error.\n')

    # Sets up the chromedriver as part of the constructor. Requires a copy of Chrome's user data folder to be located
    # in the working directory, as well as a version of chromedriver.exe that matches your web browser version.
    def chromedriver_setup(self):
        print('Initializing Webdriver')
        driver_options = webdriver.ChromeOptions()
        driver_options.add_argument('user-data-dir=' + self.cwd + '/User Data')
        chrome_service = ChromeService(self.cwd + '/chromedriver.exe')
        chrome_service.creationflags = CREATE_NO_WINDOW
        create_driver = webdriver.Chrome(options=driver_options, executable_path=self.cwd + '/chromedriver.exe',
                                         service=chrome_service)
        return create_driver

    # Closes down all pages on the chromedriver.
    def chromedriver_close(self):
        self.driver.quit()

    # Opens and returns the Excel spreadsheet for the day. If that days' spreadsheet does not exist, it
    # will create one based off the template found in the working directory.
    @staticmethod
    def create_excel_new_vendor():
        downloads_path = os.path.expanduser("~") + "/Downloads/"
        try:
            ex_workbook = load_workbook(downloads_path + "BCS_New_Vendors_" + str(datetime.date.today()) + ".xlsx")
        except FileNotFoundError:
            ex_workbook = load_workbook("BCS_template.xlsx")
        ex_worksheet = ex_workbook.worksheets[0]
        return ex_workbook, ex_worksheet

    # Opens and returns the Excel spreadsheet for the day. If that days' spreadsheet does not exist, it
    # will create one based off the template found in the working directory.
    @staticmethod
    def create_excel_existing_vendor():
        downloads_path = os.path.expanduser("~") + "/Downloads/"
        try:
            ex_workbook = load_workbook(downloads_path + "BCS_New_Vendors_" + str(datetime.date.today()) + ".xlsx")
        except FileNotFoundError:
            ex_workbook = load_workbook("BCS_template.xlsx")
        ex_worksheet = ex_workbook.worksheets[2]
        return ex_workbook, ex_worksheet

    # Saves all changes made to the Excel spreadsheet and adjust the length of cells so all information is readable.
    # If this is a newly created spreadsheet it will be named: "BCS_New_vendors_<today's_date>"
    def format_and_save_excel(self):
        print('writing to Excel')
        column_names = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P')
        downloads_path = os.path.expanduser("~") + "/Downloads/"
        index_counter = 0
        max_length = 0
        for column in self.ex_worksheet.columns:
            length_list = []
            for cell in column:
                length = len(str(cell.value))
                length_list.append(length)
                max_length = max(length_list)
            self.ex_worksheet.column_dimensions[column_names[index_counter]].width = (max_length + 2)
            index_counter += 1
        self.ex_workbook.save(downloads_path + 'BCS_New_Vendors_' + str(datetime.date.today())
                              + '.xlsx')

    # Adds the provided value parameter to the spreadsheet cell located at the provided cell_coord parameter.
    def add_to_excel(self, cell_coord, value):
        cell = self.ex_worksheet[cell_coord]
        cell.value = value

    # Iterates through the Excel spreadsheet to find and return the coordinates of the first blank cell in the
    # column identified by the target_column parameter.
    def find_blank_cell_in_column(self, target_column):
        column_coord = target_column[0].coordinate
        for row in self.ex_worksheet.iter_rows(1, 20):
            for cell in row:
                if cell.coordinate[0] == column_coord[0]:
                    if cell.value is None:
                        return cell.coordinate
                    else:
                        continue
                else:
                    continue

    # Takes info saved in the clipboard and saves it in memory. Both a string version of this data, and a list
    # version of this data are saved.
    def copy_and_parse_clipboard(self):
        copy_input = clipboard.paste()
        self.input_string = str(copy_input)
        # List is created by splitting the string on whitespace.
        self.input_list = self.input_string.split()

    # Opens Vcommerce using the password that the user provided to the GUI, then searches for the Vcommerce number
    # that the user provided to the GUI. GUI code is located in main_GUI.py
    def open_vcommerce(self):
        print('Loading Vcommerce')
        self.driver.get('https://vul.onelogin.com/portal')
        self.driver.maximize_window()
        # Checks to see if the user auto logs in via onelogin. If so it will access Vcommerce and if not it will log in,
        # and then access Vcommerce. I gave this a 7 second wait time to hopefully give enough time for the page to load before checking.
        try:
            print('Trying auto Login')
            time.sleep(7)
            vcommerce_button = self.driver.find_element(By.XPATH, '//a[@aria-label="Launch VCommerce - Production"]')
            vcommerce_button.click()
        except:
            print('Trying Manual Login')
            name_box_elements = self.driver.find_elements(By.ID, "username")
            if name_box_elements:
                continue_button_element = self.driver.find_element(By.XPATH, self.map.full_xpath['continue_button'])
                continue_button_element.click()
            WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH,
                                                                             self.map.full_xpath['login_pw'])))
            sleep(0.5)
            vc_login_element = self.driver.find_element(By.XPATH, self.map.full_xpath['login_pw'])
            vc_login_element.send_keys(' ')
            vc_login_element.clear()
            # The sleep function is used to slow down the program so that user input has time to be added correctly, or
            # is used to wait for a page to finish loading.
            sleep(0.5)
            vc_login_element.send_keys(self.vc_password)
            sleep(0.5)
            vc_login_element.send_keys(Keys.ENTER)
            WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH,
                                                                                self.map.full_xpath['apps_list'])))
            vcommerce_button = self.driver.find_element(By.XPATH, '//a[@aria-label="Launch VCommerce - Production"]')
            vcommerce_button.click()
        print('pass try-except')
        time.sleep(3)
        self.driver.get('https://solutions.sciquest.com/apps/Router/SupplierSearch')
        self.switch_tab_to_2()
        self.driver.close()
        self.switch_tab_to_1()
        self.driver.get(f'https://solutions.sciquest.com/apps/Router/SupplierProfileDashboard?CMMSP_SupplierID={self.vc_number}')
        WebDriverWait(self.driver, 20).until(ec.presence_of_all_elements_located((
            By.ID, "PhoenixNavLink_PHX_NAV_SupplierProfile_CorporateInfo")))
        
    # The switch_tab_to_1() method has the driver object switch to the first browser tab, which has Vcommerce loaded.
    def switch_tab_to_1(self):
        self.driver.switch_to.window(self.driver.window_handles[0])
        
    # The switch_tab_to_2() method has the driver object switch to the second browser tab, which has Vcommerce loaded.
    def switch_tab_to_2(self):
        self.driver.switch_to.window(self.driver.window_handles[1])

    # Navigates to the appropriate parts of the Vcommerce page and copies all data into the clipboard.
    # The copy_and_parse_clipboard function is later used to save this info to memory.
    def get_vc_and_contact_info(self):
        print('Extracting vendor data')
        time.sleep(1)
        vc_general_element = self.driver.find_element(By.ID, "PhoenixNavLink_PHX_NAV_SupplierProfile_CorporateInfo")
        vc_general_element.click()
        print('in general')
        time.sleep(1)
        WebDriverWait(self.driver, 20).until(ec.presence_of_all_elements_located((
            By.XPATH, self.map.full_xpath['vc_general_page_full'])))
        self.vc_service_type = self.get_service_type()
        sleep(1)
        html_element = self.driver.find_element(By.XPATH, '/html')
        html_element.send_keys(Keys.HOME)
        print('scrolled')
        time.sleep(1)
        vc_summary_element = self.driver.find_element(By.XPATH, '//*[@id="PhoenixNavLink_PHX_NAV_SupplierProfile_SupplierSummary"]')
        vc_summary_element.click()
        WebDriverWait(self.driver, 20).until(ec.presence_of_element_located((By.XPATH,
                                                                             self.map.full_xpath['vc_summary_page'])))
        time.sleep(1)
        vc_summary_page = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_summary_page_full'])
        vc_summary_page.send_keys(Keys.CONTROL + 'a')
        vc_summary_page.send_keys(Keys.CONTROL + 'c')
        vc_contact_name, vc_contact_phone, vc_contact_email = self.find_contact_info()
        name_cell = self.find_blank_cell_in_column(self.ex_worksheet['M'])
        self.add_to_excel(name_cell, vc_contact_name.title())
        email_cell = self.find_blank_cell_in_column(self.ex_worksheet['N'])
        self.add_to_excel(email_cell, vc_contact_email)
        phone_cell = self.find_blank_cell_in_column(self.ex_worksheet['O'])
        self.add_to_excel(phone_cell, vc_contact_phone)

    # The get_service_type() method checks to see if the vendor has a single service type or multiple service types.
    # If single, this service type is returned as a string. If multiple they are returned as a list of strings.
    # Of special note is the use of the self.driver.find_elements() method (note the plural find_elements). It
    # is different from the normal find_element method in that it looks for the specified locator and returns any
    # matches it finds in the form of a list of objects. This is very useful because an empty list in python is
    # considered False when used as a boolean, while a list with 1 or more objects within is considered True when
    # used as a boolean. This allows the method to be used to check for the presence of a potential web element,
    # and return True or False depending on if it is found or not.
    def get_service_type(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(1)
        multi_service = self.driver.find_elements(By.XPATH, '//*[@id="CustElement_11297_popover"]')
        print(multi_service)
        if len(multi_service) == 1:
            service_type_element = self.driver.find_element(By.XPATH, '//*[@id="CustElement_11297_popover"]')
            service_type_element.click()
            service_type_text = self.driver.find_element(By.XPATH, '/html/body/div[15]/div/div[2]/div/div/div').text
            service_type_list = service_type_text.split('\n')
            return service_type_list
        else:      
            single_service_type = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_service']).text
            return single_service_type

    # Checks for a single vendor contact or multiple contacts, then saves their name, email, and phone to memory.
    # This part got a little messy when updating to work on othther machines, Vcommerce can have a slightly different
    # HTML layout depending on who is logging in. The following function essentially tests different HTML path's and 
    # goes with the one that works. If I had more time I would make this more elegant.
    def find_contact_info(self):
        print('Finding vendor contact info')
        time.sleep(1)
        xpath_choice_bool = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact'])
        xpath_choice_bool2 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact2'])
        xpath_choice_bool3 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact3'])
        xpath_choice_bool4 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact4'])
        xpath_choice_bool5 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact5'])
        xpath_choice_bool6 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact6'])
        xpath_choice_bool7 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact7'])
        xpath_choice_bool8 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_single_contact8'])
        if xpath_choice_bool:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact'])
        elif xpath_choice_bool2:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact2'])
        elif xpath_choice_bool3:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact3'])
        elif xpath_choice_bool4:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact4'])
        elif xpath_choice_bool5:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact5'])  
        elif xpath_choice_bool6:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact6']) 
        elif xpath_choice_bool7:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact7'])
        elif xpath_choice_bool8:
            vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_single_contact8'])
        else:
            vendor_contact_element_test = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_multi_contact'])
            vendor_contact_element_test2 = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_multi_contact2'])
            if vendor_contact_element_test:
                vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_multi_contact'])
            elif vendor_contact_element_test2:
                vendor_contact_element = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_multi_contact2'])
            else:
                vc_name, vc_phone, vc_email = 'N/A'
                return vc_name, vc_phone, vc_email

        vendor_contact_element.click()
        time.sleep(1)
        vendor_popup_check = self.driver.find_elements(By.XPATH, self.map.full_xpath['vc_contact_popup2'])
        if vendor_popup_check:
            vendor_popup_text = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_contact_popup2']).text
            vendor_popup_list = vendor_popup_text.split()
        else:
            vendor_popup_text = self.driver.find_element(By.XPATH, self.map.full_xpath['vc_contact_popup']).text
            vendor_popup_list = vendor_popup_text.split()

        vendor_popup_list = vendor_popup_text.split()
        vc_name = vendor_popup_list[0] + ' ' + vendor_popup_list[1]
        if vendor_popup_list[2] == '+1':
            vc_phone = vendor_popup_list[3]
        else:
            vc_phone = vendor_popup_list[-2]
        vc_email = vendor_popup_list[-1]
        return vc_name, vc_phone, vc_email

    # Navigates to the certificate of insurance link in Vcommerce and clicks to download it.
    def find_and_download_cert(self):
        print('Downloading Certificate')
        compliance_link = self.driver.find_element(By.XPATH, self.map.full_xpath['legal_and_compliance_link'])
        compliance_link.click()
        sleep(0.5)
        insurance_link = self.driver.find_element(By.XPATH, self.map.full_xpath['insurance_custom_link'])
        insurance_link.click()
        time.sleep(5)
        cert_available = self.driver.find_elements(By.XPATH, self.map.full_xpath['cert_download'])
        cert_available_alternate = self.driver.find_elements(By.XPATH, self.map.full_xpath['cert_download2'])
        if cert_available:
            cert_download = self.driver.find_element(By.XPATH, self.map.full_xpath['cert_download'])
            cert_download.click()
            self.wait_for_download()
        elif cert_available_alternate:
            cert_download = self.driver.find_element(By.XPATH, self.map.full_xpath['cert_download2'])
            cert_download.click()
            self.wait_for_download()
        else:
            print('No certificate provided')

    # Checks the download folder to see if there is a file being downloaded. Files currently being downloaded
    # by chrome have the suffix .crdownload, so this method checks the downloads folder and only proceeds when
    # a .crdownload file is not found.
    @staticmethod
    def wait_for_download():
        download_path = os.path.expanduser("~") + "/Downloads/"
        download_wait = True
        while download_wait:
            time.sleep(1)
            download_wait = False
            for item in os.listdir(download_path):
                if item.endswith('.crdownload'):
                    download_wait = True
                    break

    # Searches the saved list attribute and identifies the supplier id number. This method then calls the
    # add_to_excel() method to save it to the spreadsheet. If there is no supplier id found it will return
    # a boolean to tell the find_vc_number() method to use the vcommerce number in place of the supplier id number.
    def find_supplier_id(self):
        print('Finding supplier ID')
        for item in self.input_list:
            if item == 'Number':
                temp_index = self.input_list.index(item)
                if self.input_list[temp_index - 1] == 'Supplier':
                    supplier_id = self.input_list[temp_index + 1]
                    if supplier_id != 'No':
                        sleep(0.5)
                        blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['A'])
                        self.add_to_excel(blank_cell, supplier_id)
                        return False
                    else:
                        return True
                else:
                    self.input_list.remove(item)
                    continue
        return True

    # Searches the saved list attribute and identifies the vcommerce number. It then calls the add_to_excel()
    # method to save it to the spreadsheet. If the find_supplier_id() method returns True, this will use the
    # vcommerce number in place of the supplier ID number as well.
    def find_vc_number(self, no_id):
        vcommerce_number_found = False
        vcommerce_number = 0
        print('Finding vcommerce number')
        for item in self.input_list:
            if item == 'ID':
                temp_index = self.input_list.index(item)
                vcommerce_number = self.input_list[temp_index + 1]
                sleep(0.5)
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['P'])
                self.add_to_excel(blank_cell, vcommerce_number)
                vcommerce_number_found = True
                break
        if no_id and vcommerce_number_found:
            sleep(0.5)
            blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['A'])
            self.add_to_excel(blank_cell, vcommerce_number)
        else:
            pass

    # Takes in the vendors name a parameter, and correctly formats the grammar and capitalization, as the names are
    # usually all caps in the vendor registration system. The function then returns the correctly formatted name.
    # Also identifies if a DBA name is being used and formats this correctly as well.
    @staticmethod
    def vendor_name_format(vendor_name_unformatted_arg):
        if vendor_name_unformatted_arg[0:20] == 'Doing Business As : ':
            temp_vendor_name = vendor_name_unformatted_arg[20:]
            vendor_name_formatted1 = temp_vendor_name.title()
        else:
            vendor_name_formatted1 = vendor_name_unformatted_arg.title()
        if (vendor_name_formatted1[-4:]) == 'Llc\r':
            vendor_name_formatted2 = (vendor_name_formatted1[:-4] + 'LLC')
        else:
            vendor_name_formatted2 = vendor_name_formatted1[:-1]
        vendor_name_split = vendor_name_formatted2.split()
        vendor_name_formatted3 = ''
        for item in vendor_name_split:
            if item not in ['Of', 'Is', 'The']:
                vendor_name_formatted3 += item + ' '
            else:
                vendor_name_formatted3 += item.lower() + ' '
        if vendor_name_formatted3[0:3] == 'the':
            vendor_name_formatted3 = 'The' + vendor_name_formatted3[3:]
        vendor_name_formatted3 = vendor_name_formatted3[:-1]
        if vendor_name_formatted3[-3:] == 'Inc':
            vendor_name_formatted3 = vendor_name_formatted3 + '.'
        return vendor_name_formatted3
    
    # Searches the saved list attribute to identify the name of the vendor. It will then pass that name through the 
    # vendor_name_format() method to correct its formatting and then will call the add_to_excel() method to
    # add this info to the spreadsheet.
    def find_vendor_name(self):
        print('Finding vendor name')
        input_list_split = self.input_string.split('\n')
        for item in input_list_split:
            if item == 'Registration Status\r':
                temp_index = input_list_split.index(item)
                if input_list_split[temp_index - 2] == 'Supplier Number\r':
                    vendor_name_unformatted = input_list_split[temp_index - 4]
                    vendor_name = self.vendor_name_format(vendor_name_unformatted)
                    sleep(0.5)
                    blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['B'])
                    self.add_to_excel(blank_cell, vendor_name)
                else:
                    vendor_name_unformatted = input_list_split[temp_index - 2]
                    vendor_name = self.vendor_name_format(vendor_name_unformatted)
                    sleep(0.5)
                    blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['B'])
                    self.add_to_excel(blank_cell, vendor_name)

    # Identifies and saves the vendors DBA names from the extracted Vcommerce Information.
    def find_vendor_dba(self):
        print('Finding vendor DBA')
        input_list_split = self.input_string.split('\n')
        for item in input_list_split:
            if item == 'Registration Status\r':
                temp_index = input_list_split.index(item)
                if input_list_split[temp_index - 2] == 'Supplier Number\r':
                    vendor_dba = self.vendor_name_format(input_list_split[temp_index - 3])
                    if vendor_dba == '':
                        blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['C'])
                        self.add_to_excel(blank_cell, 'N/A')
                    else:
                        blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['C'])
                        self.add_to_excel(blank_cell, vendor_dba)
                elif input_list_split[temp_index - 1] != '\r':
                    print(repr(input_list_split[temp_index - 1]))
                    vendor_dba = self.vendor_name_format(input_list_split[temp_index - 1])
                    blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['C'])
                    self.add_to_excel(blank_cell, vendor_dba)
                else:
                    blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['C'])
                    self.add_to_excel(blank_cell, 'N/A')

    # Formats the vendor address provided in the address parameter. This will correct issues with spacing,
    # capitalization, and identify PO Box addresses.
    @staticmethod
    def format_address(address):
        formatted_address = ''
        space = False
        first_letter = True
        for letter in address:
            if letter == ' ':
                space = True
                formatted_address += letter
            elif letter != ' ' and space:
                formatted_address += letter.upper()
                space = False
            elif letter != ' ' and not space and first_letter:
                formatted_address += letter.upper()
                first_letter = False
            elif letter != ' ' and not space and not first_letter:
                formatted_address += letter.lower()
        if formatted_address[0:7] == 'P O Box':
            formatted_address = formatted_address[0:1] + formatted_address[2:]
        return formatted_address

    # Identifies the string containing the vendors City, State, and zip code from the vendor address.
    # Will also identify which zip code format is being used (5 digit or 9 digit) and save the correct
    # information accordingly.
    @staticmethod
    def find_city_state_zip(csz):
        city = ''
        state_and_zip = ''
        print('state and zip:' + state_and_zip)
        comma = False
        for letter in csz:
            # All text before the comma is the city, all text after is the state and zip code.
            if letter == ',':
                comma = True
                continue
            if not comma:
                city += letter
            if comma:
                state_and_zip += letter
        # Checks to see if there is a dash after the 5th digit of the zip code in order to determine its format.
        if state_and_zip[-6] == '-':
            zip_code = state_and_zip[-11:-6]
            state = state_and_zip[1:-12]
        else:
            zip_code = state_and_zip[-6:-1]
            state = state_and_zip[1:-7]
        return city, state, zip_code

    # Adds the vendor address, city, state, and zip code to the Excel spreadsheet.
    def enter_vendor_address(self, formatted_address, city, state, zip_code):
        address_cell = self.find_blank_cell_in_column(self.ex_worksheet['D'])
        self.add_to_excel(address_cell, formatted_address)
        city_cell = self.find_blank_cell_in_column(self.ex_worksheet['E'])
        self.add_to_excel(city_cell, city.title())
        state_cell = self.find_blank_cell_in_column(self.ex_worksheet['F'])
        self.add_to_excel(state_cell, state.title())
        zip_cell = self.find_blank_cell_in_column(self.ex_worksheet['G'])
        self.add_to_excel(zip_cell, zip_code)

    # Identifies the string containing the vendor address from the extracted info.
    def find_vendor_address(self):
        print('Finding vendor address')
        input_list_split = self.input_string.split('\n')
        for item in input_list_split:
            if item[0:7] == 'Address':
                temp_index = input_list_split.index(item)
                address = (input_list_split[temp_index])[8:-1]
                if input_list_split[temp_index + 3] == 'United States\r':
                    csz_arg = input_list_split[temp_index + 2]
                else:
                    csz_arg = input_list_split[temp_index + 1]
                formatted_address = self.format_address(address)
                # if formatted_address[0] not in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'P', 'p']:
                #    raise Exception('Address does not begin with a number and is not a PO box.')
                print('csz arg:' + csz_arg)
                city, state, zip_code = self.find_city_state_zip(csz_arg)
                if state == 'Vi':
                    state = 'U.S. Virgin Islands'
                self.enter_vendor_address(formatted_address, city, state, zip_code)
                return state

    # Adds the vendor service type to the Excel spreadsheet.
    def enter_service_type(self):
        print('Finding service type')
        # used if only one service type.
        if type(self.vc_service_type) == str:
            blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['L'])
            self.add_to_excel(blank_cell, self.vc_service_type)
        # used if multiple service types.
        else:
            multi_service_string = ''
            for item in self.vc_service_type:
                multi_service_string += item + ', '
            blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['L'])
            self.add_to_excel(blank_cell, multi_service_string[:-2])

    # Identifies the risk tier based off of the service type. All references are found in the vendor_info module.
    def enter_vendor_tier(self):
        print('Finding vendor tier')
        if type(self.vc_service_type) == str:
            vendor_tier = self.vendor_info.service_type_info[self.vc_service_type]
            if vendor_tier == 1:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 1')
            elif vendor_tier == 2:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 2')  
            elif vendor_tier == 3:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 3')
            elif vendor_tier == 4:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Marine')           
        else:
            tier_list = []
            for item in self.vc_service_type:
                service_tier = self.vendor_info.service_type_info[item]
                tier_list.append(service_tier)
            test_num = tier_list[0]
            for num in tier_list:
                if num < test_num and num != 4:
                    test_num = num
                elif num >= test_num and num != 4:
                    continue
                elif num == 4:
                    test_num = 4
                    break
            if test_num == 1:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 1')
            elif test_num == 2:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 2')
            elif test_num == 3:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Tier 3')
            elif test_num == 4:
                blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['H'])
                self.add_to_excel(blank_cell, 'Marine')

    # Identifies the company division this vendor falls under, based on the vendor's State address.
    # All references are found in the vendor_info module.
    def find_division(self, state):
        print('Finding Division')
        post_info = self.vendor_info.post_info[state]
        if type(post_info) == list:
            company_division = post_info[1]
            blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['J'])
            self.add_to_excel(blank_cell, company_division)
        else:
            raise Exception('Company Division for this state is not available in the vendor_info object')

    # identifies the POST contact, based on the vendors State address.
    # All references are found in the vendor_info module.
    def find_post_owner(self, state):
        print('Finding POST owner')
        post_owner_tuple = self.vendor_info.post_info[state]
        if type(post_owner_tuple) == list:
            post_owner = post_owner_tuple[2]
            blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['K'])
            self.add_to_excel(blank_cell, post_owner)
        else:
            raise Exception('POST Owner for this state is not available in the vendor_info object')

    # Enters 'Vulcan Materials Company' into the correct cell. This will be added the same way
    # for all vendors, regardless of their State address.
    def enter_vmc(self):
        blank_cell = self.find_blank_cell_in_column(self.ex_worksheet['I'])
        self.add_to_excel(blank_cell, 'Vulcan Materials Company')
            
    # Extract instructions. Calls all previous methods in the correct order and passes parameters between them as
    # needed. This will Successfully log into Vcommerce, search the vendor, pull its info, format the info, and add it
    # to a spreadsheet. This will also download the vendors provided certificates of insurance.
    def extract_new(self):
            self.ex_workbook, self.ex_worksheet = self.create_excel_new_vendor()
            self.verify_user_input()
            self.open_vcommerce()
            self.get_vc_and_contact_info()
            self.find_and_download_cert()
            self.copy_and_parse_clipboard()
            no_id = self.find_supplier_id()
            self.find_vc_number(no_id)
            self.find_vendor_name()
            self.find_vendor_dba()
            state = self.find_vendor_address()
            self.enter_service_type()
            self.enter_vendor_tier()
            self.find_division(state)
            self.find_post_owner(state)
            self.enter_vmc()
            self.format_and_save_excel()
            # clears the information from the clipboard once the extract is done.
            clipboard.copy('')
            print('Extract Completed')

    def extract_existing(self):
            self.ex_workbook, self.ex_worksheet = self.create_excel_existing_vendor()
            self.verify_user_input()
            self.open_vcommerce()
            self.get_vc_and_contact_info()
            self.find_and_download_cert()
            self.copy_and_parse_clipboard()
            no_id = self.find_supplier_id()
            self.find_vc_number(no_id)
            self.find_vendor_name()
            self.find_vendor_dba()
            state = self.find_vendor_address()
            self.enter_service_type()
            self.enter_vendor_tier()
            self.find_division(state)
            self.find_post_owner(state)
            self.enter_vmc()
            self.format_and_save_excel()
            # clears the information from the clipboard once the extract is done.
            clipboard.copy('')
            print('Extract Completed')
