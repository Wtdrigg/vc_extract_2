"""
This module is used to read the cumulative vendor report provided to us by BCS and identify which vendors have been added since the last time the program was run.
Note that the I/O was a recent change in my last week at Vulcan, as this previously saved into a database being hosted localy on my machine using Flask. It saves to 
.pkl files from the pickle module, but some variable names may still allude to the prior way this was saved.
"""

import os
import pickle
import glob
import csv
from xlsxwriter.workbook import Workbook as WriteWorkbook
from openpyxl import load_workbook

class Approve:

    def __init__(self):
        self.approval_csv_workbook, self.approval_csv_worksheet = self.load_approval_csv()
        self.approved_vendor_list = []
        self.past_approvals_obj = None

    # Loads the extracted report from BCS by looking in the users downloads folder for the first file that has a .csv extension.
    # Compares the list of vendors from the .csv file to the list pulled from the .pkl file.
    # Any vendors on the csv that are not on the pkl will be displayed on the GUI.
    @staticmethod
    def load_approval_csv():
        downloaded_csv_files = glob.glob(os.path.expanduser("~") + "/Downloads/*.csv")
        most_recent_csv = max(downloaded_csv_files, key=os.path.getctime)
        temp_workbook = WriteWorkbook(most_recent_csv[:-4] + ".xlsx")
        temp_worksheet = temp_workbook.add_worksheet()
        with open(most_recent_csv, 'r') as f:
            reader = csv.reader((line.replace('\0', '-') for line in f))
            for count, row in enumerate(reader):
                for count2, column in enumerate(row):
                    temp_worksheet.write(count, count2, column)
        temp_workbook.close()
        try:
            ex_workbook_csv = load_workbook(most_recent_csv[:-4] + ".xlsx")
        except FileNotFoundError:
            print('NO CSV FOUND')
            return
        ex_worksheet_csv = ex_workbook_csv.worksheets[0]
        return ex_workbook_csv, ex_worksheet_csv

    # Loads a list of previously approved vendors from past_approvals.pkl
    # the pkl file is just a serialized python object saved in the department 
    # network drive.
    def get_approval_list(self):
        self.past_approvals_obj = self.load_past_approvals()
        for json_item in self.past_approvals_obj.past_approvals:
            vendor_name = (json_item['vendor_name'])
            self.approved_vendor_list.append(vendor_name)

    # Adds the vendors that were not found in the .pkl file to the .pkl file and clears the GUI.
    def confirm_approval(self):
        for item in self.approval_csv_worksheet:
            if item[1].value not in self.approved_vendor_list:
                if item[1].row != 1:
                    id_num = (len(self.approved_vendor_list) + 1)
                    new_dict = {"id":id_num, "vendor_name":item[1].value}
                    print(new_dict)
                    self.past_approvals_obj.past_approvals.append(new_dict)
                    self.save_past_approvals(self.past_approvals_obj)
                    self.approved_vendor_list = []
                    self.get_approval_list()

    # loads the serialized pkl file and converts it to a python object of type PastApprovals.
    def load_past_approvals(self):
        with open('J:/VC_Extract_Files/past_approvals.pkl', 'rb') as f:
            past_approvals = pickle.load(f)
            return past_approvals
    
    # saves a PastApprovals object as a serialized .pkl file.
    def save_past_approvals(self, obj):
        with open('J:/VC_Extract_Files/past_approvals.pkl', 'wb') as f:
            pickle.dump(obj, f)
        
class PastApprovals:

    def __init__(self):
        self.past_approvals = []


