import os
from vendor_requests import RequestsVendors
import glob
import csv
from xlsxwriter.workbook import Workbook as WriteWorkbook
from openpyxl import Workbook
from openpyxl import load_workbook


class Approval:

    def __init__(self):
        self.approval_csv_workbook, self.approval_csv_worksheet = self.load_approval_csv()
        self.approved_vendor_list = []
        self.requests = RequestsVendors()


    @staticmethod
    def load_approval_csv():
        downloaded_csv_files = glob.glob(os.path.expanduser("~") + "/Downloads/*.csv")
        most_recent_csv = max(downloaded_csv_files, key=os.path.getctime)
        temp_workbook = WriteWorkbook(most_recent_csv[:-4] + ".xlsx")
        temp_worksheet = temp_workbook.add_worksheet()
        with open(most_recent_csv, 'r') as f:
            reader = csv.reader((line.replace('\0', '-') for line in f))
            for count, row in enumerate(reader):
                for count2, column in enumerate(row):
                    temp_worksheet.write(count, count2, column)
        temp_workbook.close()
        try:
            ex_workbook_csv = load_workbook(most_recent_csv[:-4] + ".xlsx")
        except FileNotFoundError:
            print('NO CSV FOUND')
            return
        ex_worksheet_csv = ex_workbook_csv.worksheets[0]
        return ex_workbook_csv, ex_worksheet_csv


    def get_approval_list(self):
        past_approvals = self.requests.api_get_request()
        for json_item in past_approvals:
            vendor_name = (json_item['vendor_name'])
            self.approved_vendor_list.append(vendor_name)


    def confirm_approval(self):
        for item in self.approval_csv_worksheet:
             if item[1].value not in self.approved_vendor_list:
                if item[1].row != 1:
                    id = (len(self.approved_vendor_list) + 1)
                    self.requests.api_put_request(id, item[1].value)
                    self.approved_vendor_list = []
                    self.get_approval_list()


