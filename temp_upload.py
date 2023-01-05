import requests
import os
import glob
import csv
import time
from xlsxwriter.workbook import Workbook as WriteWorkbook
from openpyxl import Workbook
from openpyxl import load_workbook


class Approval:

    def __init__(self):
        self.approval_workbook, self.approval_worksheet = self.setup_vendor_list()
        self.approval_csv_workbook, self.approval_csv_worksheet = self.load_approval_csv()
        self.approved_vendor_list = []

    @staticmethod
    def setup_vendor_list():
        cwd_path = os.getcwd()
        file_name = cwd_path + "/BCS_Approved_New_Hires.xlsx"
        try:
            ex_workbook = load_workbook(cwd_path + "/BCS_Approved_New_Hires.xlsx")
        except FileNotFoundError:
            ex_workbook = Workbook()
            ex_workbook.save(file_name)
        ex_worksheet = ex_workbook.worksheets[0]
        return ex_workbook, ex_worksheet

    @staticmethod
    def load_approval_csv():
        downloaded_csv_files = glob.glob(os.path.expanduser("~") + "/Downloads/*.csv")
        most_recent_csv = max(downloaded_csv_files, key=os.path.getctime)
        temp_workbook = WriteWorkbook(most_recent_csv[:-4] + ".xlsx")
        temp_worksheet = temp_workbook.add_worksheet()
        with open(most_recent_csv, 'r') as f:
            reader = csv.reader((line.replace('\0', '-') for line in f))
            for count, row in enumerate(reader):
                for count2, column in enumerate(row):
                    temp_worksheet.write(count, count2, column)
        temp_workbook.close()
        try:
            ex_workbook_csv = load_workbook(most_recent_csv[:-4] + ".xlsx")
        except FileNotFoundError:
            print('NO CSV FOUND')
            return
        ex_worksheet_csv = ex_workbook_csv.worksheets[0]
        return ex_workbook_csv, ex_worksheet_csv

    def get_approval_list(self):
        for row in self.approval_worksheet:
            vendor_tuple = (row[0].value, row[1].value)
            self.approved_vendor_list.append(vendor_tuple)

    def confirm_approval(self):
        previous_approval_names = []
        for row in self.approval_worksheet:
            approved_vendor_name = row[1].value
            previous_approval_names.append(approved_vendor_name)
        for item in self.approved_vendor_list:
            if item[1] not in previous_approval_names:
                target_cell_vc_number = self.find_blank_cell_in_column(self.approval_worksheet['A'])
                self.add_to_excel(target_cell_vc_number, item[0])
                target_cell_vendor_name = self.find_blank_cell_in_column(self.approval_worksheet['B'])
                self.add_to_excel(target_cell_vendor_name, item[1])

    # Iterates through the Excel spreadsheet to find and return the coordinates of the first blank cell in the
    # column identified by the target_column parameter.
    def find_blank_cell_in_column(self, target_column):
        column_coord = target_column[0].coordinate
        for row in self.approval_worksheet.iter_rows(1, 500):
            for cell in row:
                if cell.coordinate[0] == column_coord[0]:
                    if cell.value is None:
                        return cell.coordinate
                    else:
                        continue
                else:
                    continue

    # Adds the provided value parameter to the spreadsheet cell located at the provided cell_coord parameter.
    def add_to_excel(self, cell_coord, value):
        cell = self.approval_worksheet[cell_coord]
        cell.value = value

    # Saves all changes made to the Excel spreadsheet and adjust the length of cells so all information is readable.
    # If this is a newly created spreadsheet it will be named: "BCS_New_vendors_<today's_date>"
    def format_and_save_excel(self):
        column_names = ('A', 'B')
        cwd_path = os.getcwd()
        file_name = cwd_path + "/BCS_Approved_New_Hires.xlsx"
        index_counter = 0
        max_length = 0
        for column in self.approval_worksheet.columns:
            length_list = []
            for cell in column:
                length = len(str(cell.value))
                length_list.append(length)
                max_length = max(length_list)
            self.approval_worksheet.column_dimensions[column_names[index_counter]].width = (max_length + 2)
            index_counter += 1
        self.approval_workbook.save(file_name)


if __name__ == '__main__':
    x = Approval()
    x.get_approval_list()
    url = 'http://wtdrigg.pythonanywhere.com/vendors'

    for count, item in enumerate(x.approved_vendor_list):
        temp_id = count + 1
        #print(f'ID = {temp_id}\nVC Number = {item[0]}\nName = {item[1]}\n')
        result = requests.put(url, {'id': temp_id, 'vendor_name': item[1]})
        print(result.text)
        time.sleep(1)
